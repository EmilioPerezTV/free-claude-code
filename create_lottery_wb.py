import openpyxl
from openpyxl.workbook import Workbook

def create_lottery_workbook(filename="analisis_loteria_avanzado.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Analisis_Avanzado"

    # Encabezados mejorados
    headers = ["ID", "Numero Ganador", "Fecha", "Frecuencia Relativa", "Media Movil (3)", "Desviacion Estandar", "Variacion vs Media"]
    ws.append(headers)

    # Simulación de datos históricos (30 sorteos)
    for i in range(2, 32):
        row = i - 1
        ws[f"A{i}"] = row
        ws[f"B{i}"] = (i * 3) % 45 + 1  # Simulación más realista
        ws[f"C{i}"] = f"2026-06-{i if i < 30 else 30:02d}"

    # Fórmulas dinámicas
    ws["H1"] = 30 # Total de sorteos

    # Columna D: Frecuencia (proporción sobre total histórico)
    for i in range(2, 32):
        ws[f"D{i}"] = f"=B{i}/SUM($B$2:$B$31)"

    # Columna E: Media Móvil (3)
    for i in range(4, 32):
        ws[f"E{i}"] = f"=AVERAGE(B{i-2}:B{i})"

    # Columna F: Desviación Estándar (muestra)
    for i in range(4, 32):
        ws[f"F{i}"] = f"=STDEV.S(B{i-2}:B{i})"

    # Columna G: Variación (Diferencia porcentual con la media)
    for i in range(4, 32):
        ws[f"G{i}"] = f"=(B{i}-E{i})/E{i}"

    wb.save(filename)
    print(f"Libro de trabajo '{filename}' creado con éxito.")

if __name__ == "__main__":
    create_lottery_workbook()
